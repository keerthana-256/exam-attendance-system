from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl

from django.contrib.auth.models import User
from .models import (
    Year, Branch, Section, Hall, Exam,
    Student, Attendance, Invigilator, HallAssignment
)
from .forms import LoginForm




def invigilator_login(request):
    if request.method == "POST":
        form = LoginForm(request.POST)

        if form.is_valid():
            user = authenticate(
                request,
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password']
            )

            if user and Invigilator.objects.filter(user=user).exists():
                login(request, user)
                return redirect('invigilator_dashboard')

            return render(request, 'exam/invigilator_login.html', {
                'form': form,
                'error': "Invalid credentials"
            })
    else:
        form = LoginForm()

    return render(request, 'exam/invigilator_login.html', {'form': form})


def admin_login(request):
    if request.method == "POST":
        user = authenticate(
            request,
            username=request.POST.get("username"),
            password=request.POST.get("password")
        )

        if user and user.is_staff:
            login(request, user)
            return redirect('admin_dashboard')

        return render(request, 'exam/admin_login.html', {
            'error': "Invalid Admin Credentials"
        })

    return render(request, 'exam/admin_login.html')




@login_required
def invigilator_dashboard(request):
    invigilator = get_object_or_404(Invigilator, user=request.user)

    assignments = HallAssignment.objects.filter(invigilator=invigilator)

    return render(request, 'exam/dashboard.html', {
        'assignments': assignments
    })



@login_required
def takeAttendance(request, hall_no):
    hall = get_object_or_404(Hall, hall_no=hall_no)


    if not request.user.is_staff:
        invigilator = get_object_or_404(Invigilator, user=request.user)

        assigned = HallAssignment.objects.filter(
            invigilator=invigilator,
            hall=hall
        ).exists()

        if not assigned:
            return redirect('invigilator_dashboard')

    students_qs = Student.objects.filter(hall=hall)

    if not students_qs.exists():
        return render(request, 'exam/takeAttendance.html', {
            'students': [],
            'hall': hall,
            'is_locked': False
        })

    exam = students_qs.first().exam

    if exam.start_time:
        exam_datetime = datetime.combine(exam.date, exam.start_time)
        exam_datetime = timezone.make_aware(exam_datetime)
        lock_time = exam_datetime + timedelta(minutes=30)
        is_locked = timezone.now() > lock_time
    else:
        is_locked = False

    if is_locked:
        return render(request, 'exam/attendance_locked.html', {
            'hall': hall,
            'exam': exam
        })

    existing = Attendance.objects.filter(student__in=students_qs)

    attendance_map = {
        record.student_id: record.status
        for record in existing
    }

    students = []
    for s in students_qs:
        students.append({
            "id": s.id,
            "reg_no": s.reg_no,
            "name": s.name,
            "exam": s.exam.subject,
            "status": attendance_map.get(s.id, "Present")
        })

    if request.method == "POST":
        for s in students_qs:
            status = request.POST.get(s.reg_no, "Present")

            Attendance.objects.update_or_create(
                student=s,
                defaults={'status': status}
            )

        return redirect('success')

    return render(request, 'exam/takeAttendance.html', {
        'students': students,
        'hall': hall,
        'is_locked': is_locked
    })


def success(request):
    return render(request, 'exam/success.html')


@staff_member_required
def admin_dashboard(request):
    exams = Exam.objects.all()
    selected_exam_id = request.GET.get("exam")

    records = Attendance.objects.select_related(
        'student', 'student__exam', 'student__hall'
    )

    if selected_exam_id:
        records = records.filter(student__exam_id=selected_exam_id)

    return render(request, 'exam/admin_dashboard.html', {
        'records': records,
        'exams': exams,
        'selected_exam_id': selected_exam_id
    })


@staff_member_required
def upload_students(request):
    if request.method == "POST":
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            reg_no, name, year_name, branch_name, section_name, hall_name, subject, date, session = row

            if isinstance(date, datetime):
                date = date.date()

            year, _ = Year.objects.get_or_create(year_name=str(year_name).strip())
            branch, _ = Branch.objects.get_or_create(branch_name=str(branch_name).strip())

            section, _ = Section.objects.get_or_create(
                section_name=str(section_name).strip(),
                year=year,
                branch=branch
            )

            hall, _ = Hall.objects.get_or_create(hall_no=str(hall_name).strip())

            exam, _ = Exam.objects.get_or_create(
                subject=str(subject).strip(),
                date=date,
                session=str(session).strip()
            )

            Student.objects.update_or_create(
                reg_no=str(reg_no).strip(),
                defaults={
                    'name': name,
                    'year': year,
                    'branch': branch,
                    'section': section,
                    'hall': hall,
                    'exam': exam
                }
            )

        return redirect('admin_dashboard')

    return render(request, 'exam/upload.html')


@staff_member_required
def upload_invigilators(request):
    if request.method == "POST":
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            username, hall_name, date, session = row

            if isinstance(date, datetime):
                date = date.date()

            username = str(username).strip()

            try:
                user = User.objects.get(username=username)
            except User.DoesNotExist:
                continue

            invigilator, _ = Invigilator.objects.get_or_create(
                user=user,
                defaults={'name': user.first_name or user.username}
            )

         
            hall_obj, _ = Hall.objects.get_or_create(
                hall_no=str(hall_name).strip()
            )

            HallAssignment.objects.update_or_create(
                invigilator=invigilator,
                hall=hall_obj,
                date=date,
                session=str(session).strip()
            )

        return redirect('admin_dashboard')

    return render(request, 'exam/upload_invigilators.html')

@staff_member_required
def section_wise_absentees(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)

    sections = Section.objects.all()
    section_data = {}

    for section in sections:
        absentees = Attendance.objects.filter(
            student__section=section,
            student__exam=exam,
            status="Absent"
        ).select_related('student', 'student__hall')

        if absentees.exists():
            section_data[section] = [
                record.student for record in absentees
            ]

    return render(request, 'exam/section_absentees.html', {
        'section_data': section_data,
        'exam': exam
    })

staff_member_required
def export_attendance_excel(request, exam_id):

    records = Attendance.objects.filter(
        student__exam_id=exam_id
    ).select_related('student').order_by('student__reg_no')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.append([
        "Register No",
        "Student Name",
        "Status"
    ])

    for record in records:
        ws.append([
            record.student.reg_no,
            record.student.name,
            record.status
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename=attendance_{exam_id}.xlsx'
    )

    wb.save(response)
    return response


def create_admin(request):
    from django.contrib.auth.models import User
    from django.http import HttpResponse

    if not User.objects.filter(username="Keerthana").exists():
        User.objects.create_superuser("Keerthana", "keerthanapasuparthi@gmail.com", "keerthana@256")
    return HttpResponse("Admin created")

from django.http import HttpResponse
from django.core.management import call_command

def run_migrations(request):
    call_command('migrate')
    return HttpResponse("Migrations done")